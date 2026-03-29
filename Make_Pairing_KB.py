import openpyxl
from openpyxl.utils import get_column_letter
from os import path, remove

kbFile = '.\\datasets\\kb.pl'
rulesFile='.\\datasets\\Regole_Abbinamento.txt'

if path.exists(kbFile):
    remove(kbFile)

viniFile = '.\\datasets\\vini_mercadini.xlsx'
piattiFile = '.\\datasets\\cucina_italiana_mercadini.xlsx'

print("Creazione Knowledge Base...")

##########################################################################################################
# Caricamento fatti relativi ai vini
# Usa l'intestazione per mappare colonne
vino_prolog_lines = []

workbook = openpyxl.load_workbook(viniFile)
worksheet = workbook.active  

header_row = 1  # Riga 1 contiene header "# | Vino | Regione..."

col_mapping = {}
for col in range(2, 13):  # Colonne B-L (2-12)
    header = str(worksheet.cell(row=header_row, column=col).value).strip()
    col_mapping[header] = col

# Colonne necessarie secondo ordine richiesto
required_cols = ['Vino', 'Regione / Paese', 'Tipologia', 'Effervescenza', 
                'Acidità', 'Sapidità', 'Alcoolicità', 'Tannicità', 
                'Morbidezza', 'Dolcezza', 'Persistenza']

# Processa dati dalle righe successive all'header
start_row = header_row + 1
for row_num in range(start_row, worksheet.max_row + 1):
    vino_name = str(worksheet.cell(row=row_num, column=col_mapping['Vino']).value or '').strip()
    if not vino_name or vino_name.startswith('#'):  # Salta righe vuote/commenti
        continue
    
    # Estrai valori, converti in int se numerico
    try:
        regione = str(worksheet.cell(row=row_num, column=col_mapping['Regione / Paese']).value or '').strip().replace("'", "\\'")
        tipologia = str(worksheet.cell(row=row_num, column=col_mapping['Tipologia']).value or '').strip().lower()
        eff = int(worksheet.cell(row=row_num, column=col_mapping['Effervescenza']).value or 1)
        acid = int(worksheet.cell(row=row_num, column=col_mapping['Acidità']).value or 0)
        sap = int(worksheet.cell(row=row_num, column=col_mapping['Sapidità']).value or 0)
        alco = int(worksheet.cell(row=row_num, column=col_mapping['Alcoolicità']).value or 0)
        tann = int(worksheet.cell(row=row_num, column=col_mapping['Tannicità']).value or 0)
        morb = int(worksheet.cell(row=row_num, column=col_mapping['Morbidezza']).value or 0)
        dolce = int(worksheet.cell(row=row_num, column=col_mapping['Dolcezza']).value or 0)
        pers = int(worksheet.cell(row=row_num, column=col_mapping['Persistenza']).value or 0)
        
        # Formato Prolog: vino('Nome', 'Regione', tipologia, eff, acid, sap, alco, tann, morb, dolce, pers).
        # Escape apostrofi nel nome vino
        vino_escaped = vino_name.replace("'", "\\'")
        prolog_line = f"vino('{vino_escaped}', '{regione}', '{tipologia}', {eff}, {acid}, {sap}, {alco}, {tann}, {morb}, {dolce}, {pers})."
        vino_prolog_lines.append(prolog_line)
        
    except (ValueError, KeyError, TypeError) as e:
        print(f"Errore riga {row_num}: {e}")
        continue

print(f"Generati {len(vino_prolog_lines)} fatti Prolog in KB.pl sui vini")

##########################################################################################################
# Caricamento fatti relativi ai piatti
# Usa l'intestazione per mappare colonne
piatti_prolog_lines = []

workbook = openpyxl.load_workbook(piattiFile)
worksheet = workbook.active  
header_row = 1  

col_mapping = {}
for col in range(1, 13):  # Colonne A-L (1-12)
    header = str(worksheet.cell(row=header_row, column=col).value).strip()
    col_mapping[header] = col

# Colonne necessarie secondo ordine richiesto
required_cols = ['Piatto', 'Categoria', 'Tend. Dolce', 'Grassezza', 
                'Untuosità', 'Succulenza', 'Speziatura', 'Sapidità', 
                'Tend. Amarognola', 'Tend. Acida', 'Dolcezza', 'Persistenza']

# Processa dati dalle righe successive all'header
start_row = header_row + 1
for row_num in range(start_row, worksheet.max_row + 1):
    piatto_name = str(worksheet.cell(row=row_num, column=col_mapping['Piatto']).value or '').strip().replace("'", "\\'")
    if not piatto_name or piatto_name.startswith('#'):  # Salta righe vuote/commenti
        continue
    
    # Estrai valori, converti in int se numerico
    try:
        categoria = str(worksheet.cell(row=row_num, column=col_mapping['Categoria']).value or '').strip().replace("'", "\\'")
        tend_dolce = int(worksheet.cell(row=row_num, column=col_mapping['Tend. Dolce']).value or 0)
        grass = int(worksheet.cell(row=row_num, column=col_mapping['Grassezza']).value or 0)
        untuos = int(worksheet.cell(row=row_num, column=col_mapping['Untuosità']).value or 0)        
        succ = int(worksheet.cell(row=row_num, column=col_mapping['Succulenza']).value or 0)
        spez = int(worksheet.cell(row=row_num, column=col_mapping['Speziatura']).value or 0)
        sap = int(worksheet.cell(row=row_num, column=col_mapping['Sapidità']).value or 0)
        tend_amar = int(worksheet.cell(row=row_num, column=col_mapping['Tend. Amarognola']).value or 0)
        tend_acid = int(worksheet.cell(row=row_num, column=col_mapping['Tend. Acida']).value or 0)
        dolce = int(worksheet.cell(row=row_num, column=col_mapping['Dolcezza']).value or 0)
        pers = int(worksheet.cell(row=row_num, column=col_mapping['Persistenza']).value or 0)
        
        # Formato Prolog: piatto(nome, categoria, tend_dolce, grassezza, untuosita, succulenza, speziatura, sapidita, tend_amara, tend_acida,dolcezza, persistenza).        
        prolog_line = f"piatto('{piatto_name}', '{categoria}', {tend_dolce}, {grass}, {untuos}, {succ}, {spez}, {sap}, {tend_amar}, {tend_acid}, {dolce}, {pers})."
        piatti_prolog_lines.append(prolog_line)
        
    except (ValueError, KeyError, TypeError) as e:
        print(f"Errore riga {row_num}: {e}")
        continue

print(f"Generati {len(piatti_prolog_lines)} fatti Prolog in KB.pl sui piatti")

# Scrivi file KB.pl
with open(kbFile, 'w', encoding='utf-8') as f:
    f.write(':- encoding(utf8).\n')
    f.write('%===================================================\n')
    f.write('% KB Abbinamento Cibo - Vino secondo Mercadini\n')
    f.write('%===================================================\n')
    for line in vino_prolog_lines:
        f.write(line + '\n')
    for line in piatti_prolog_lines:
        f.write(line + '\n')
    
    with open(rulesFile, 'r', encoding='utf-8') as regole_file:            
        for line in regole_file:
            f.write(line)

print("Generazione del file KB.pl terminata!")
exit()

